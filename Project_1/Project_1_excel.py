import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        cell.value = cell.value * 0.9

    wb.save(filename)

    values = Reference(sheet,
                       min_row=2,
                        max_row=sheet.max_row,
                       min_col=3,
                       max_col=3)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'd2')

    wb.save(filename)


process_workbook("transactions.xlsx")